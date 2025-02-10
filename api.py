import time
import pandas as pd
import base64
from openpyxl import load_workbook
import json
from flask import Flask, request, jsonify
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import os

app = Flask(__name__)

# Base URL (change this to your hosted link)
base_url = "https://qr-code-ashen-omega.vercel.app/"

# Define column indices
start_row = 2
name_col, id_col, domain_col, duration_col = 1, 2, 3, 4
task1_col, task2_col, task3_col, task4_col, link_col = 5, 6, 7, 8, 10

def generate_link(name, student_id, domain, duration, tasks):
    """Generate an encoded link for the intern data."""
    data = {
        "name": name,
        "id": student_id,
        "domain": domain,
        "duration": duration,
        "tasks": [task for task in tasks if task],
    }
    json_data = json.dumps(data)
    encoded_data = base64.urlsafe_b64encode(json_data.encode()).decode()
    return f"{base_url}?data={encoded_data}"

@app.route("/generate-link", methods=["POST"])
def generate_link_api():
    """API endpoint to generate a link dynamically."""
    try:
        data = request.json
        link = generate_link(
            data.get("name"),
            data.get("id"),
            data.get("domain"),
            data.get("duration"),
            data.get("tasks", [])
        )
        return jsonify({"link": link})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/process-excel", methods=["POST"])
def process_excel():
    """API to process an uploaded Excel file and generate links."""
    try:
        file = request.files['file']
        if not file:
            return jsonify({"error": "No file provided"}), 400

        wb = load_workbook(file)
        sheet = wb.active

        for row in range(start_row, sheet.max_row + 1):
            if sheet.cell(row=row, column=link_col).value:
                continue
            
            name = sheet.cell(row=row, column=name_col).value
            student_id = sheet.cell(row=row, column=id_col).value
            domain = sheet.cell(row=row, column=domain_col).value
            duration = sheet.cell(row=row, column=duration_col).value
            tasks = [
                sheet.cell(row=row, column=task1_col).value,
                sheet.cell(row=row, column=task2_col).value,
                sheet.cell(row=row, column=task3_col).value,
                sheet.cell(row=row, column=task4_col).value,
            ]
            
            link = generate_link(name, student_id, domain, duration, tasks)
            sheet.cell(row=row, column=link_col).value = link

        return jsonify({"message": "Links generated successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
